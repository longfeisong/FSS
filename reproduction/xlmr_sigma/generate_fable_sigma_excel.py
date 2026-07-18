#!/usr/bin/env python3
"""Generate the report-ready workbook for the complete FABLE + SIGMA run."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[2]
RESULT_PATH = ROOT / "reproduction/xlmr_sigma/results/fable_sigma_full.json"
OUTPUT = ROOT / "FABLE_SIGMA完整链路实验汇报.xlsx"

NAVY = "17365D"
BLUE = "4472C4"
LIGHT_BLUE = "D9EAF7"
GREEN = "70AD47"
LIGHT_GREEN = "E2F0D9"
ORANGE = "ED7D31"
LIGHT_ORANGE = "FCE4D6"
RED = "C00000"
LIGHT_RED = "F4CCCC"
GRAY = "666666"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="B7B7B7")


def title(ws, text: str, subtitle: str, end_col: int = 8) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=end_col)
    c = ws.cell(1, 1, text)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.font = Font(color=WHITE, bold=True, size=20)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=end_col)
    s = ws.cell(3, 1, subtitle)
    s.font = Font(color=GRAY, italic=True, size=10)
    s.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28


def header(cells) -> None:
    for cell in cells:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def table(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in range(start_row + 1, end_row + 1):
        if (row - start_row) % 2 == 0:
            for col in range(start_col, end_col + 1):
                ws.cell(row, col).fill = PatternFill("solid", fgColor="F5F8FC")


def widths(ws, values: dict[str, float]) -> None:
    for column, value in values.items():
        ws.column_dimensions[column].width = value


def metrics(result: dict) -> dict[str, float]:
    lookup = result["lookup"]
    bridge = result["fable_to_sigma_mask"]
    sigma = result["sigma"]
    bridge_total_bytes = bridge["bytes_sent_per_party"] * 2
    bridge_wall_ms = max(bridge["elapsed_us_p0"], bridge["elapsed_us_p1"]) / 1000
    sigma_wall_ms = max(
        sigma["online_elapsed_us_p0_batch"], sigma["online_elapsed_us_p1_batch"]
    ) / 1000
    total_online_bytes = (
        lookup["total_sent_bytes"]
        + bridge_total_bytes
        + sigma["online_communication_bytes_per_party_batch"]
    )
    total_online_ms = lookup["wall_ms"] + bridge_wall_ms + sigma_wall_ms
    return {
        "bridge_total_bytes": bridge_total_bytes,
        "bridge_wall_ms": bridge_wall_ms,
        "sigma_wall_ms": sigma_wall_ms,
        "total_online_bytes": total_online_bytes,
        "total_online_ms": total_online_ms,
    }


def make_summary(wb: Workbook, r: dict, m: dict[str, float]) -> None:
    ws = wb.active
    ws.title = "汇报总览"
    ws.sheet_view.showGridLines = False
    title(
        ws,
        "FABLE + SIGMA 完整链路阶段结果",
        "FacebookAI/xlm-roberta-base｜4×128 tokens｜512×768 私密 Embedding｜真实 XLM-R Encoder Layer 0",
    )

    ws.merge_cells("A5:H7")
    c = ws["A5"]
    c.value = (
        "阶段结论：已完成并验证“私密 Token IDs + 私密 250,002×768 Embedding 表 → FABLE → "
        "50-bit 算术份额 → SIGMA 掩码输入 → 真实 XLM-R Encoder Block”。"
        "393,216 个查表值零错误，四条序列的 P0/P1 输出全部一致，sequence 0 与零掩码 correctness oracle 完全一致。"
    )
    c.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    c.font = Font(color="375623", bold=True, size=13)
    c.alignment = Alignment(vertical="center", wrap_text=True)
    c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    cols = ["阶段", "工作量/配置", "耗时", "通信/存储", "正确性", "关键意义"]
    for i, value in enumerate(cols, 1):
        ws.cell(9, i, value)
    header(ws[9][:6])
    lookup = r["lookup"]
    bridge = r["fable_to_sigma_mask"]
    sigma = r["sigma"]
    rows = [
        [
            "FABLE 私密查表",
            "512 unique IDs；250,002×768 表；24 chunks",
            f"{lookup['wall_ms'] / 1000:,.3f} s",
            f"双方发送合计 {lookup['total_sent_bytes']:,} B",
            f"{lookup['values_checked']:,} values；{lookup['mismatches']} mismatch",
            "同时保护查询 ID 与 Embedding 表；当前 24 次分片尚未优化",
        ],
        [
            "FABLE→SIGMA 转换",
            "393,216 个 50-bit ring elements",
            f"P0/P1：{bridge['elapsed_us_p0']/1000:,.3f}/{bridge['elapsed_us_p1']/1000:,.3f} ms",
            f"每方发送 {bridge['bytes_sent_per_party']:,} B",
            "masked input 相同；0 mismatch",
            "不公开重构 Embedding，将算术份额安全转换为 x+r",
        ],
        [
            "SIGMA 离线预处理",
            "1 layer × 4 sequences；密钥不可复用",
            "未纳入在线时间",
            f"每方 batch 密钥 {sigma['offline_key_bytes_per_party_batch']:,} B",
            "四条序列均生成独立密钥",
            "体现 FSS 方案的离线存储成本",
        ],
        [
            "SIGMA 在线推理",
            "XLM-R Layer 0；4×128×768；12 heads",
            f"P0/P1 batch：{sigma['online_elapsed_us_p0_batch']/1e6:,.3f}/{sigma['online_elapsed_us_p1_batch']/1e6:,.3f} s",
            f"每方 batch {sigma['online_communication_bytes_per_party_batch']:,} B",
            "四条序列 P0=P1；seq0=oracle",
            "证明 FABLE 输出能进入真实权重 SIGMA Block",
        ],
        [
            "顺序组合总计（推导）",
            "Lookup + share→mask + 4 次 SIGMA online",
            f"约 {m['total_online_ms']/1000:,.3f} s",
            f"约 {m['total_online_bytes']:,} B",
            "各阶段均通过",
            "用于展示当前实现规模；不是优化后的最终性能",
        ],
    ]
    for row, values in enumerate(rows, 10):
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
        ws.row_dimensions[row].height = 50
    table(ws, 9, 9 + len(rows), 1, 6)

    ws.merge_cells("A17:H20")
    warning = ws["A17"]
    warning.value = (
        "汇报边界：本次 SIGMA 时间不能作为正式性能数字，因为运行前 A100-2 utilization=99%。"
        "FABLE 查表按原生 512-bit 输出接口重复 24 个 chunk，通信和时间属于 correctness-first 实现，不是优化结果。"
        "当前语义是 word embeddings 直接进入一个 Encoder Block，尚未加入 XLM-R position embeddings、embedding LayerNorm，"
        "也尚未运行完整 12 层。因此可以汇报“安全链路已跑通”，不能汇报“完整 XLM-R 推理性能”。"
    )
    warning.fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
    warning.font = Font(color="9C5700", bold=True)
    warning.alignment = Alignment(vertical="center", wrap_text=True)
    warning.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # Log-scale phase charts. Helper data stay visible for auditability.
    ws["J9"] = "在线阶段"
    ws["K9"] = "时间（ms）"
    ws["L9"] = "通信（B）"
    phase_rows = [
        ("FABLE Lookup", lookup["wall_ms"], lookup["total_sent_bytes"]),
        ("Share→Mask", m["bridge_wall_ms"], m["bridge_total_bytes"]),
        ("SIGMA Online", m["sigma_wall_ms"], sigma["online_communication_bytes_per_party_batch"]),
    ]
    for row, values in enumerate(phase_rows, 10):
        for col, value in enumerate(values, 10):
            ws.cell(row, col, value)
    header(ws[9][9:12])
    table(ws, 9, 12, 10, 12)
    for col, anchor, chart_title, axis_title in (
        (11, "J15", "各在线阶段耗时（对数坐标）", "ms"),
        (12, "J31", "各在线阶段通信（对数坐标）", "Bytes"),
    ):
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = chart_title
        chart.y_axis.title = axis_title
        chart.y_axis.scaling.logBase = 10
        chart.height = 8
        chart.width = 14
        chart.add_data(Reference(ws, min_col=col, min_row=9, max_row=12), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=10, min_row=10, max_row=12))
        chart.legend = None
        ws.add_chart(chart, anchor)

    widths(ws, {"A": 23, "B": 34, "C": 25, "D": 30, "E": 29, "F": 43, "G": 3, "H": 3, "J": 22, "K": 20, "L": 22})
    ws.freeze_panes = "A10"
    ws.print_area = "A1:H20"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def make_configuration(wb: Workbook, r: dict) -> None:
    ws = wb.create_sheet("配置与安全")
    ws.sheet_view.showGridLines = False
    title(ws, "实验配置与安全接口", "说明实际运行了什么，以及各方能够看到什么", 7)
    headers = ["类别", "配置项", "数值", "P0 可见", "P1 可见", "Dealer 可见", "关键意义"]
    for i, value in enumerate(headers, 1):
        ws.cell(5, i, value)
    header(ws[5])
    sigma = r["sigma"]
    rows = [
        ["模型", "模型及 revision", f"{r['model']}\n{r['model_revision']}", "模型方表/掩码参数", "公共掩码参数", "参数 mask", "固定模型来源，保证可复现"],
        ["模型", "Encoder 深度", "已运行 Layer 0（1/12 layers）", "—", "—", "—", "当前是单 Block 链路验证，不是完整模型"],
        ["模型", "Hidden / heads / FFN", "768 / 12 / 3072", "—", "—", "—", "与 SIGMA bert-base 形状一致"],
        ["输入", "Sequence batch", "4×128 tokens = 512 IDs", "不见明文 IDs", "私密 IDs", "不见 IDs", "对齐 FABLE 作者 512-query batch"],
        ["查表", "Embedding table", "250,002×768，int16 scale 12", "私密表", "不见表", "不见表", "选择大规模 LUT 场景"],
        ["FABLE 输出", "算术份额", "x0+x1=x mod 2^50", "x0", "x1", "不见 x", "查表结果不公开重构"],
        ["SIGMA 输入", "掩码形式", "public x+r；dealer retains r", "x+r 与 r0", "x+r 与 r1", "完整 r", "安全连接两种协议表示"],
        ["SIGMA 参数", "权重形式", "dealer mask / evaluator masked weights", "公共 masked weights", "公共 masked weights", "weight mask", "没有把真实权重错误地明文加载给双方"],
        ["数值", "定点与 ring", f"scale={sigma['fixed_point_scale']}；bitwidth={sigma['ring_bitwidth']}", "—", "—", "—", "与 SIGMA bert-base 算术配置一致"],
        ["硬件", "GPU", "P0=A100-7；P1=A100-2", "A100-7", "A100-2", "A100-7", "A100-2 当时 utilization 99%，计时无效"],
    ]
    for row, values in enumerate(rows, 6):
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
        ws.row_dimensions[row].height = 48
    table(ws, 5, 5 + len(rows), 1, 7)
    widths(ws, {"A": 14, "B": 25, "C": 36, "D": 25, "E": 25, "F": 25, "G": 42})
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:G{5 + len(rows)}"


def make_details(wb: Workbook, r: dict, m: dict[str, float]) -> None:
    ws = wb.create_sheet("指标明细")
    ws.sheet_view.showGridLines = False
    title(ws, "结果指标及关键意义", "区分正确性、安全性、在线性能、离线成本和结果可靠性", 7)
    headers = ["类别", "指标", "数值", "单位", "关键意义", "当前结论", "是否可作正式性能"]
    for i, value in enumerate(headers, 1):
        ws.cell(5, i, value)
    header(ws[5])
    lookup, bridge, sigma = r["lookup"], r["fable_to_sigma_mask"], r["sigma"]
    rows = [
        ["正确性", "FABLE values checked", lookup["values_checked"], "values", "验证整张查询输出，不只检查样例", "393,216 个值全部匹配", "是"],
        ["正确性", "FABLE mismatch", lookup["mismatches"], "values", "查表正确性的硬门槛", "0 mismatch", "是"],
        ["正确性", "Share→mask mismatch", bridge["mismatches"], "values", "证明 FABLE 输出能正确转换为 SIGMA 表示", "0 mismatch", "是"],
        ["正确性", "P0/P1 outputs", "4/4 identical", "sequences", "证明双方对安全推理结果达成一致", "全部一致", "是"],
        ["正确性", "Zero-mask oracle", "sequence 0 identical", "sequence", "排除“双方一致但共同算错”的风险", "逐字节一致", "是（正确性）"],
        ["在线性能", "FABLE wall", lookup["wall_ms"] / 1000, "s", "当前端到端的主要时间瓶颈", "约 12.315 分钟", "否：24 chunks 未优化"],
        ["在线性能", "SIGMA online batch P0/P1", f"{sigma['online_elapsed_us_p0_batch']/1e6:.3f} / {sigma['online_elapsed_us_p1_batch']/1e6:.3f}", "s", "4 条序列的 Transformer 在线延迟", "协议已跑通", "否：GPU utilization 99%"],
        ["在线性能", "顺序组合总时间", m["total_online_ms"] / 1000, "s", "展示当前实现的整体数量级", "约 12.406 分钟", "否：仅阶段性结果"],
        ["通信", "FABLE total sent", lookup["total_sent_bytes"], "B", "衡量查表网络成本", "当前最大通信瓶颈", "否：分片查询未优化"],
        ["通信", "Share→mask total sent", m["bridge_total_bytes"], "B", "衡量协议衔接自身开销", "相比 FABLE Lookup 很小", "可作为实现指标"],
        ["通信", "SIGMA online / party / batch", sigma["online_communication_bytes_per_party_batch"], "B", "衡量 Transformer 在线网络成本", "4 条序列合计", "可报告通信量"],
        ["离线成本", "SIGMA key / party / sequence", sigma["offline_key_bytes_per_party_per_sequence"], "B", "FSS 密钥不能跨输入复用", "每条序列约 1.507 GB", "是"],
        ["离线成本", "SIGMA key / party / batch", sigma["offline_key_bytes_per_party_batch"], "B", "影响磁盘和预处理带宽", "4 条序列约 6.028 GB", "是"],
        ["可靠性", "GPU utilization", "A100-2=99%；A100-7=0%", "%", "决定延迟是否能进入正式平均值", "timing_reliable=false", "否"],
        ["语义边界", "完整 XLM-R", "未完成", "—", "决定能否报告模型级效果", "缺 position embedding、embedding LN 和其余 11 层", "否"],
    ]
    for row, values in enumerate(rows, 6):
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
        ws.row_dimensions[row].height = 46
    table(ws, 5, 5 + len(rows), 1, 7)
    widths(ws, {"A": 14, "B": 31, "C": 31, "D": 15, "E": 42, "F": 36, "G": 24})
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:G{5 + len(rows)}"


def flatten(prefix: str, value, rows: list[list]) -> None:
    if isinstance(value, dict):
        for key, child in value.items():
            flatten(f"{prefix}.{key}" if prefix else key, child, rows)
    elif isinstance(value, list):
        rows.append([prefix, json.dumps(value, ensure_ascii=False), "array", "fable_sigma_full.json"])
    else:
        rows.append([prefix, value, type(value).__name__, "fable_sigma_full.json"])


def make_raw(wb: Workbook, r: dict, m: dict[str, float]) -> None:
    ws = wb.create_sheet("原始数据")
    ws.sheet_view.showGridLines = False
    title(ws, "原始结果与推导值", "原始 JSON 字段完整展开，便于复核", 5)
    for i, value in enumerate(["JSON 字段", "数值", "类型", "来源", "备注"], 1):
        ws.cell(5, i, value)
    header(ws[5])
    rows: list[list] = []
    flatten("", r, rows)
    rows.extend(
        [
            ["derived.bridge_total_bytes", m["bridge_total_bytes"], "number", "2×bytes_sent_per_party", "两方发送合计"],
            ["derived.bridge_wall_ms", m["bridge_wall_ms"], "number", "max(P0,P1 elapsed)", "组合 wall 口径"],
            ["derived.sigma_wall_ms", m["sigma_wall_ms"], "number", "max(P0,P1 batch elapsed)", "组合 wall 口径"],
            ["derived.total_online_bytes", m["total_online_bytes"], "number", "Lookup+bridge+SIGMA", "顺序组合推导"],
            ["derived.total_online_ms", m["total_online_ms"], "number", "Lookup+bridge+SIGMA", "顺序组合推导"],
        ]
    )
    for row, values in enumerate(rows, 6):
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
    table(ws, 5, 5 + len(rows), 1, 5)
    widths(ws, {"A": 55, "B": 52, "C": 14, "D": 29, "E": 30})
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:E{5 + len(rows)}"


def make_wording(wb: Workbook, r: dict) -> None:
    ws = wb.create_sheet("汇报口径")
    ws.sheet_view.showGridLines = False
    title(ws, "建议汇报口径", "可以汇报的成果、必须附带的边界和下一阶段指标", 6)
    sections = [
        (
            "推荐主结论",
            LIGHT_GREEN,
            [
                "已经实现 FABLE 与 SIGMA 的安全协议衔接，FABLE 输出不会以明文形式重构。",
                "完整验证了 512×768 个 Embedding 输出，393,216 个定点值零 mismatch。",
                "FABLE 输出成功进入真实 XLM-R Encoder Layer 0，四条序列的双方输出全部一致。",
                "sequence 0 与零掩码 SIGMA correctness oracle 逐字节一致，排除了共同错误。",
            ],
        ),
        (
            "必须附带的限制",
            LIGHT_ORANGE,
            [
                "当前 FABLE 使用 24 个 32 维 chunk，重复了查询/setup 工作，因此 738.922 秒和 191.5 GB 不是优化后性能。",
                "SIGMA 运行时 A100-2 utilization=99%，约 5.4 秒 batch 时间只用于正确性，不进入正式性能平均值。",
                "当前只运行一个 XLM-R Encoder Block，尚未运行全部 12 层。",
                "尚未加入 position embeddings 和 embedding LayerNorm，因此不能称为完整 XLM-R 语义推理。",
            ],
        ),
        (
            "下一阶段正式对比",
            LIGHT_BLUE,
            [
                "固定同一个 baseline，只比较 Lookup 实现差异，不再同时改变模型、输入和隐私目标。",
                "正式统计 Lookup setup/online、share conversion、SIGMA dealer/online、端到端通信和峰值内存。",
                "GPU utilization≤30% 后重复至少 5 次，报告平均值±标准差。",
                "先优化 FABLE 24 chunks 的查询与 setup 复用，再与 baseline 计算 speedup 和 communication reduction。",
            ],
        ),
        (
            "不要使用的表述",
            LIGHT_RED,
            [
                "不要说“FABLE+SIGMA 已经比 baseline 更快”——baseline 尚未按同口径实现和测量。",
                "不要把当前 738.922 秒称为 FABLE 的最终性能——这是 correctness-first 分片实现。",
                "不要说“完整 XLM-R 已跑通”——目前是完整私密 Embedding 加一个真实 Encoder Block。",
            ],
        ),
    ]
    row = 5
    for section, color, bullets in sections:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row, 1, section)
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(bold=True, size=13)
        c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        row += 1
        for bullet in bullets:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            c = ws.cell(row, 1, f"• {bullet}")
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            ws.row_dimensions[row].height = 36
            row += 1
        row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(
        row,
        1,
        f"结果提交：f2219db（FABLE+SIGMA 实现）｜报表生成：{datetime.now().astimezone().isoformat(timespec='seconds')}",
    ).font = Font(color=GRAY, italic=True)
    widths(ws, {"A": 27, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18})


def validate(path: Path, r: dict) -> None:
    wb = load_workbook(path, data_only=False)
    expected = ["汇报总览", "配置与安全", "指标明细", "原始数据", "汇报口径"]
    if wb.sheetnames != expected:
        raise AssertionError(wb.sheetnames)
    if len(wb["汇报总览"]._charts) != 2:
        raise AssertionError("expected two phase charts")
    if wb["汇报总览"]["A10"].value != "FABLE 私密查表":
        raise AssertionError("summary table missing")
    if r["lookup"]["mismatches"] != 0 or r["fable_to_sigma_mask"]["mismatches"] != 0:
        raise AssertionError("cannot report a failed correctness run")
    if path.stat().st_size < 12_000:
        raise AssertionError("workbook unexpectedly small")


def main() -> None:
    r = json.loads(RESULT_PATH.read_text(encoding="utf-8"))
    m = metrics(r)
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    make_summary(wb, r, m)
    make_configuration(wb, r)
    make_details(wb, r, m)
    make_raw(wb, r, m)
    make_wording(wb, r)
    wb.properties.title = "FABLE + SIGMA 完整链路实验汇报"
    wb.properties.subject = "XLM-R private Embedding to SIGMA Encoder Block"
    wb.properties.creator = "FSS experiment reproduction"
    wb.save(OUTPUT)
    validate(OUTPUT, r)
    print(OUTPUT)
    print(f"derived_total_online_ms={m['total_online_ms']:.3f}")
    print(f"derived_total_online_bytes={m['total_online_bytes']}")
    print(f"bytes={OUTPUT.stat().st_size}")


if __name__ == "__main__":
    main()

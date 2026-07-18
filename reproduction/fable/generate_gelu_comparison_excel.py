#!/usr/bin/env python3
"""Generate the report-ready SIGMA GELU vs FABLE GELU workbook."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
OUTPUT = ROOT / "reproduction/fable/results/FABLE_GELU替换前后对比.xlsx"

# The original SIGMA run reports two BERT-Tiny blocks.  Normalize it to one
# block before comparing it with the one-block FABLE component experiment.
QUERIES = 65_536
SIGMA_BLOCKS = 2
SIGMA_TIME_TOTAL_MS = 207.078
SIGMA_COMM_TOTAL_BYTES = 3_473_408
SIGMA_TIME_MS = SIGMA_TIME_TOTAL_MS / SIGMA_BLOCKS
SIGMA_COMM_BYTES = SIGMA_COMM_TOTAL_BYTES / SIGMA_BLOCKS
FABLE_PROTOCOL_MS = 75_437
FABLE_WALL_MS = 85_000
FABLE_COMM_BYTES = 19_372_844_388
LOGICAL_LUT_ENTRIES = 256
FABLE_PADDED_ENTRIES = 65_536
FABLE_CHUNKS = 16

TIME_SLOWDOWN = FABLE_PROTOCOL_MS / SIGMA_TIME_MS
COMM_INCREASE = FABLE_COMM_BYTES / SIGMA_COMM_BYTES
PADDING_MULTIPLIER = FABLE_PADDED_ENTRIES / LOGICAL_LUT_ENTRIES
SIGMA_QPS = QUERIES / (SIGMA_TIME_MS / 1000)
FABLE_QPS = QUERIES / (FABLE_PROTOCOL_MS / 1000)
SIGMA_BYTES_PER_QUERY = SIGMA_COMM_BYTES / QUERIES
FABLE_BYTES_PER_QUERY = FABLE_COMM_BYTES / QUERIES


NAVY = "17365D"
BLUE = "4472C4"
LIGHT_BLUE = "D9EAF7"
ORANGE = "ED7D31"
LIGHT_ORANGE = "FCE4D6"
GREEN = "70AD47"
LIGHT_GREEN = "E2F0D9"
RED = "C00000"
LIGHT_RED = "F4CCCC"
GRAY = "666666"
LIGHT_GRAY = "E7E6E6"
WHITE = "FFFFFF"
THIN_GRAY = Side(style="thin", color="B7B7B7")


def style_title(ws, title: str, subtitle: str, end_column: int = 8) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=end_column)
    cell = ws.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(color=WHITE, bold=True, size=20)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=end_column)
    sub = ws.cell(3, 1, subtitle)
    sub.font = Font(color=GRAY, italic=True, size=10)
    sub.alignment = Alignment(horizontal="center")


def style_header(row) -> None:
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


def style_table(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row_idx in range(min_row + 1, max_row + 1):
        if (row_idx - min_row) % 2 == 0:
            for col_idx in range(min_col, max_col + 1):
                ws.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor="F5F8FC")


def set_widths(ws, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def make_summary(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "汇报总览"
    ws.sheet_view.showGridLines = False
    style_title(
        ws,
        "FABLE 替换 SIGMA GELU：替换前后对比",
        "统一口径：BERT-Tiny 单个 Transformer Block，65,536 次 GELU 查询",
    )

    ws.merge_cells("A5:H7")
    conclusion = ws["A5"]
    conclusion.value = (
        f"核心结论：在 256 项 GELU 修正表场景中，FABLE 没有带来性能收益。"
        f"协议时间约慢 {TIME_SLOWDOWN:,.1f}×，通信量约增加 {COMM_INCREASE:,.0f}×。"
        "根因是 FABLE BatchPIR 至少按 65,536 项处理，使逻辑表膨胀 256×，且 65,536 次查询被拆为 16 批。"
    )
    conclusion.fill = PatternFill("solid", fgColor=LIGHT_RED)
    conclusion.font = Font(color=RED, bold=True, size=13)
    conclusion.alignment = Alignment(vertical="center", wrap_text=True)
    conclusion.border = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)

    headers = ["指标", "替换前：SIGMA GELU", "替换后：FABLE GELU", "变化", "关键意义"]
    for col, value in enumerate(headers, 1):
        ws.cell(9, col, value)
    style_header(ws[9][:5])

    rows = [
        ["查询数 / Block", QUERIES, QUERIES, "相同", "工作量一致，是进行每 Block 对照的基础"],
        ["逻辑 LUT 项数", LOGICAL_LUT_ENTRIES, LOGICAL_LUT_ENTRIES, "相同", "两者计算同一个 GELU 修正函数"],
        ["协议实际 LUT 项数", LOGICAL_LUT_ENTRIES, FABLE_PADDED_ENTRIES, f"{PADDING_MULTIPLIER:,.0f}×", "FABLE 小表填充是时间和通信膨胀的主要来源"],
        ["在线/协议时间 / Block", SIGMA_TIME_MS, FABLE_PROTOCOL_MS, f"慢 {TIME_SLOWDOWN:,.1f}×", "决定在线推理延迟；越低越好"],
        ["通信量 / Block", SIGMA_COMM_BYTES, FABLE_COMM_BYTES, f"高 {COMM_INCREASE:,.0f}×", "决定跨机器带宽成本；越低越好"],
        ["查询吞吐量（query/s）", SIGMA_QPS, FABLE_QPS, f"降至 {FABLE_QPS / SIGMA_QPS:.4%}", "反映单位时间可完成的 GELU 查询数量；越高越好"],
        ["通信量 / query（B）", SIGMA_BYTES_PER_QUERY, FABLE_BYTES_PER_QUERY, f"高 {COMM_INCREASE:,.0f}×", "排除查询数量影响后的单次查询网络成本"],
        ["正确性", "零输入 SIGMA smoke", "16 chunks 全部零错误", "均通过各自检查", "证明实现可运行，但两种正确性输入并不完全等价"],
    ]
    for row_idx, values in enumerate(rows, 10):
        for col_idx, value in enumerate(values, 1):
            ws.cell(row_idx, col_idx, value)
    style_table(ws, 9, 9 + len(rows), 1, 5)

    for row_idx in (13, 14, 15, 16):
        ws.cell(row_idx, 2).number_format = '#,##0.000'
        ws.cell(row_idx, 3).number_format = '#,##0.000'
    for row_idx in (10, 11, 12):
        ws.cell(row_idx, 2).number_format = '#,##0'
        ws.cell(row_idx, 3).number_format = '#,##0'
    for row_idx in range(10, 18):
        ws.row_dimensions[row_idx].height = 34

    ws.merge_cells("A20:H22")
    caveat = ws["A20"]
    caveat.value = (
        "实验边界：该结果用于说明数量级差异，不是严格公平的端到端推理对比。"
        "SIGMA 数值来自零输入/零权重 smoke run 的两个 Block，并已归一化为每 Block；"
        "FABLE 是随机有效 GELU 索引的独立组件实验，尚未接入 live GPU tensor。"
        "共享服务器负载也使时间只能标记为 provisional。"
    )
    caveat.fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
    caveat.font = Font(color="9C5700", bold=True)
    caveat.alignment = Alignment(vertical="center", wrap_text=True)
    caveat.border = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)

    # Helper data for a report-friendly ratio chart.
    ws["J9"] = "开销指标"
    ws["K9"] = "替换后 / 替换前"
    ws["J10"] = "协议时间"
    ws["K10"] = TIME_SLOWDOWN
    ws["J11"] = "通信量"
    ws["K11"] = COMM_INCREASE
    ws["J12"] = "协议 LUT 项数"
    ws["K12"] = PADDING_MULTIPLIER
    style_header(ws[9][9:11])
    style_table(ws, 9, 12, 10, 11)
    for cell in ws["K"][9:12]:
        cell.number_format = '#,##0.0"×"'

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "替换后开销倍率（对数坐标，越低越好）"
    chart.y_axis.title = "倍率"
    chart.x_axis.title = "指标"
    chart.y_axis.scaling.logBase = 10
    chart.height = 8.5
    chart.width = 14.5
    data = Reference(ws, min_col=11, min_row=9, max_row=12)
    cats = Reference(ws, min_col=10, min_row=10, max_row=12)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, "J14")

    set_widths(ws, {"A": 25, "B": 22, "C": 22, "D": 20, "E": 48, "F": 3, "G": 3, "H": 3, "J": 22, "K": 20})
    ws.freeze_panes = "A10"
    ws.auto_filter.ref = "A9:E17"
    ws.print_area = "A1:H22"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def make_details(wb: Workbook) -> None:
    ws = wb.create_sheet("指标明细")
    ws.sheet_view.showGridLines = False
    style_title(ws, "对照指标及关键意义", "每项指标说明其工程含义、结论和可比性", 8)
    headers = ["类别", "指标", "替换前", "替换后", "变化", "关键意义", "汇报解读", "可比性说明"]
    for col, value in enumerate(headers, 1):
        ws.cell(5, col, value)
    style_header(ws[5])

    details = [
        ["工作负载", "模型结构", "BERT-Tiny：H=128, heads=2, FFN=512", "相同目标 Block", "结构一致", "控制模型规模变量", "差异主要来自 GELU 协议，而非模型尺寸", "FABLE 尚未接入完整 Block"],
        ["工作负载", "Sequence length", 128, 128, "相同", "直接决定 FFN 激活数量", "两边均对应 65,536 次 GELU", "一致"],
        ["工作负载", "GELU 查询数 / Block", QUERIES, QUERIES, "相同", "控制查询规模", "按单 Block 归一化后工作量一致", "一致"],
        ["协议参数", "逻辑 LUT 项数", LOGICAL_LUT_ENTRIES, LOGICAL_LUT_ENTRIES, "相同", "GELU 只需 8-bit 索引", "这是典型小表查找", "一致"],
        ["协议参数", "实际 LUT 项数", LOGICAL_LUT_ENTRIES, FABLE_PADDED_ENTRIES, f"{PADDING_MULTIPLIER:,.0f}×", "决定 PIR 数据库规模", "FABLE 的最小 BatchPIR 参数不适合 256 项小表", "实现约束导致差异"],
        ["协议参数", "FABLE 分批数", "不适用", f"{FABLE_CHUNKS}×4,096", "16 批", "影响重复初始化和网络往返", "进一步放大 FABLE 开销", "协议结构不同"],
        ["性能", "在线/协议时间 / Block（ms）", SIGMA_TIME_MS, FABLE_PROTOCOL_MS, f"慢 {TIME_SLOWDOWN:,.1f}×", "用户等待时间；越低越好", "当前 GELU 小表不应采用 FABLE 替换", "共享服务器，时间 provisional"],
        ["性能", "FABLE wall time（ms）", "未单独记录", FABLE_WALL_MS, "约 85 秒", "包含协议外启动、同步等开销", "反映实际运行感受", "无法与 SIGMA wall time严格配对"],
        ["性能", "查询吞吐量（query/s）", SIGMA_QPS, FABLE_QPS, f"下降 {1 - FABLE_QPS / SIGMA_QPS:.4%}", "单位时间处理能力；越高越好", "FABLE 吞吐量显著下降", "由协议时间推导"],
        ["通信", "通信量 / Block（B）", SIGMA_COMM_BYTES, FABLE_COMM_BYTES, f"高 {COMM_INCREASE:,.0f}×", "决定带宽、跨地域费用和延迟", "通信膨胀比计算膨胀更严重", "SIGMA 为两 Block总量除以2"],
        ["通信", "通信量 / query（B）", SIGMA_BYTES_PER_QUERY, FABLE_BYTES_PER_QUERY, f"高 {COMM_INCREASE:,.0f}×", "单位查询网络成本", "证明差异不是查询数造成", "由通信量和查询数推导"],
        ["正确性", "检查结果", "零输入 smoke 通过", "16 chunks 全部零错误", "均通过", "排除明显实现错误", "只能说明各自协议执行正确", "输入分布不同，不能作为精度对照"],
        ["结论边界", "端到端集成", "完整 SIGMA Block", "独立 FABLE LUT 组件", "未完成 live bridge", "决定能否称为端到端替换", "报告中应称组件级数量级对比", "不可宣称严格端到端公平比较"],
    ]
    for row_idx, values in enumerate(details, 6):
        for col_idx, value in enumerate(values, 1):
            ws.cell(row_idx, col_idx, value)
        ws.row_dimensions[row_idx].height = 48
    style_table(ws, 5, 5 + len(details), 1, 8)
    ws.conditional_formatting.add(
        f"E6:E{5 + len(details)}",
        ColorScaleRule(start_type="min", start_color=LIGHT_GREEN, mid_type="percentile", mid_value=50, mid_color="FFF2CC", end_type="max", end_color=LIGHT_RED),
    )
    set_widths(ws, {"A": 14, "B": 28, "C": 26, "D": 26, "E": 20, "F": 34, "G": 38, "H": 36})
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:H{5 + len(details)}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def make_raw_data(wb: Workbook) -> None:
    ws = wb.create_sheet("原始数据")
    ws.sheet_view.showGridLines = False
    style_title(ws, "原始数据与推导值", "保留精确数值、单位和数据来源，便于复核", 6)
    headers = ["字段", "数值", "单位", "适用对象", "来源/计算", "备注"]
    for col, value in enumerate(headers, 1):
        ws.cell(5, col, value)
    style_header(ws[5])
    raw = [
        ["SIGMA Block 数", SIGMA_BLOCKS, "blocks", "替换前", "原始 SIGMA smoke run", "原始统计覆盖两个 Block"],
        ["SIGMA GELU 总时间", SIGMA_TIME_TOTAL_MS, "ms", "替换前", "SIGMA 原始统计", "两个 Block 合计"],
        ["SIGMA GELU 时间 / Block", SIGMA_TIME_MS, "ms", "替换前", "总时间 / 2", "用于公平归一化"],
        ["SIGMA GELU 总通信", SIGMA_COMM_TOTAL_BYTES, "B", "替换前", "SIGMA 原始统计", "两个 Block 合计"],
        ["SIGMA GELU 通信 / Block", SIGMA_COMM_BYTES, "B", "替换前", "总通信 / 2", "用于公平归一化"],
        ["FABLE protocol time", FABLE_PROTOCOL_MS, "ms", "替换后", "reproduction/fable/results/runs.csv", "16 chunks 合计"],
        ["FABLE wall time", FABLE_WALL_MS, "ms", "替换后", "reproduction/fable/results/runs.csv", "包含启动和同步"],
        ["FABLE communication", FABLE_COMM_BYTES, "B", "替换后", "reproduction/fable/results/runs.csv", "一个 Block"],
        ["查询数", QUERIES, "queries/Block", "双方", "128×512", "一个 Block"],
        ["时间减速倍率", TIME_SLOWDOWN, "×", "推导", "FABLE protocol / SIGMA per Block", "越接近1越好"],
        ["通信增长倍率", COMM_INCREASE, "×", "推导", "FABLE comm / SIGMA per Block", "越接近1越好"],
        ["LUT 填充倍率", PADDING_MULTIPLIER, "×", "推导", "65,536 / 256", "FABLE BatchPIR 最小规模"],
        ["SIGMA throughput", SIGMA_QPS, "query/s", "替换前", "queries / seconds", "推导"],
        ["FABLE throughput", FABLE_QPS, "query/s", "替换后", "queries / seconds", "推导"],
        ["SIGMA bytes/query", SIGMA_BYTES_PER_QUERY, "B/query", "替换前", "comm / queries", "推导"],
        ["FABLE bytes/query", FABLE_BYTES_PER_QUERY, "B/query", "替换后", "comm / queries", "推导"],
    ]
    for row_idx, values in enumerate(raw, 6):
        for col_idx, value in enumerate(values, 1):
            ws.cell(row_idx, col_idx, value)
        ws.cell(row_idx, 2).number_format = '#,##0.000'
    style_table(ws, 5, 5 + len(raw), 1, 6)
    set_widths(ws, {"A": 30, "B": 22, "C": 18, "D": 16, "E": 42, "F": 32})
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:F{5 + len(raw)}"


def make_reporting_notes(wb: Workbook) -> None:
    ws = wb.create_sheet("汇报口径")
    ws.sheet_view.showGridLines = False
    style_title(ws, "建议汇报口径", "区分可以下结论的内容与不能过度外推的内容", 6)
    sections = [
        (
            "推荐主结论",
            LIGHT_GREEN,
            [
                f"在 BERT-Tiny GELU 的 256 项小表场景中，FABLE 协议时间约慢 {TIME_SLOWDOWN:,.1f} 倍。",
                f"FABLE 通信量约增加 {COMM_INCREASE:,.0f} 倍，是更突出的系统瓶颈。",
                "主要原因不是 FABLE 查询错误，而是其 BatchPIR 最小数据库规模使 256 项表填充到 65,536 项。",
                "因此研究路线转向 XLM-R Embedding 等大规模查表，而不继续用 FABLE 替换小型 GELU LUT。",
            ],
        ),
        (
            "必须附带的实验边界",
            LIGHT_ORANGE,
            [
                "SIGMA 原始数据覆盖两个 Block，本表已除以 2 归一化到每 Block。",
                "SIGMA 是零输入/零权重 smoke run；FABLE 是随机有效索引的组件实验。",
                "FABLE 尚未在该 GELU 实验中连接 live GPU tensor，因此这是组件级数量级对比。",
                "服务器为共享环境，时间结果应标记 provisional；通信数量级不受 GPU 竞争直接影响。",
            ],
        ),
        (
            "不要使用的表述",
            LIGHT_RED,
            [
                "不要说“FABLE 在任何查表上都比 SIGMA 慢”。本实验只证明它不适合 256 项 GELU 小表。",
                "不要说“完成了严格端到端替换”。当时只完成了 FABLE LUT 组件验证。",
                "不要把 207.078 ms 直接与 75,437 ms 比；前者是两个 Block，必须先除以 2。",
                "不要把两种正确性测试描述为相同输入上的模型精度比较。",
            ],
        ),
    ]
    row = 5
    for title, color, bullets in sections:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        header = ws.cell(row, 1, title)
        header.fill = PatternFill("solid", fgColor=color)
        header.font = Font(bold=True, size=13)
        header.alignment = Alignment(vertical="center")
        header.border = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
        row += 1
        for bullet in bullets:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row, 1, f"• {bullet}")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
            ws.row_dimensions[row].height = 34
            row += 1
        row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row, 1, f"生成时间：{datetime.now().astimezone().isoformat(timespec='seconds')}  |  Git commit：fa8de58（GELU 实验）").font = Font(color=GRAY, italic=True)
    set_widths(ws, {"A": 25, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18})
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def validate(path: Path) -> None:
    book = load_workbook(path, data_only=False)
    expected = ["汇报总览", "指标明细", "原始数据", "汇报口径"]
    if book.sheetnames != expected:
        raise AssertionError(book.sheetnames)
    if book["汇报总览"]["B13"].value != SIGMA_TIME_MS:
        raise AssertionError("SIGMA per-block normalization missing")
    if book["汇报总览"]["C13"].value != FABLE_PROTOCOL_MS:
        raise AssertionError("FABLE protocol time missing")
    if book["汇报总览"]["B14"].value != SIGMA_COMM_BYTES:
        raise AssertionError("SIGMA per-block communication missing")
    if book["汇报总览"]["C14"].value != FABLE_COMM_BYTES:
        raise AssertionError("FABLE communication missing")
    if path.stat().st_size < 10_000:
        raise AssertionError("workbook is unexpectedly small")


def main() -> None:
    workbook = Workbook()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    make_summary(workbook)
    make_details(workbook)
    make_raw_data(workbook)
    make_reporting_notes(workbook)
    workbook.properties.title = "FABLE 替换 SIGMA GELU 前后对比"
    workbook.properties.subject = "BERT-Tiny 单 Transformer Block GELU 查表实验"
    workbook.properties.creator = "FSS experiment reproduction"
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    validate(OUTPUT)
    print(OUTPUT)
    print(f"time_slowdown={TIME_SLOWDOWN:.6f}x")
    print(f"communication_increase={COMM_INCREASE:.6f}x")
    print(f"bytes={OUTPUT.stat().st_size}")


if __name__ == "__main__":
    main()
